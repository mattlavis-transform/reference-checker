import json
import re
import os
from docx import Document
from openpyxl import Workbook, load_workbook

import classes.globals as g
from classes.preference import Preference
from classes.preference_error import PreferenceError
from classes.database import Database


class PreferenceChecker(object):
    def __init__(self, origin, origin_object = None):
        self.origin = origin
        self.preference_errors = []
        if origin_object is None:
            self.get_preference_config()
        else:
            self.country = origin_object["country"]
            self.source = os.path.join(g.app.preferences_source, origin_object["source"])
            self.dest = os.path.join(g.app.preferences_dest, origin + ".json")
            a = 1
            
        print("Checking preferences for", self.origin)
        # self.start_log()
        self.get_preference_data_from_db()
        self.get_preferences_from_excel()
        self.compare_preferences()
        self.validate_mismatches()
        self.check_hierarchy()
        self.expand_differences()
        self.write_preference_errors()
        
    def validate_mismatches(self):
        for preference_error in self.preference_errors:
            if preference_error["reason"] == "Mismatched duty":
                if self.fta_preferences_dict[preference_error["commodity_code"]]["instance_count"] > 1:
                    preference_error["ignore"] = True
            a = 1

    def get_preference_config(self):
        with open(g.app.preferences_config) as jsonFile:
            jsonObject = json.load(jsonFile)
            obj = jsonObject[self.origin]
            self.country = obj["country"]
            self.source = os.path.join(g.app.preferences_source, obj["source"])
            self.dest = os.path.join(g.app.preferences_dest, self.origin + ".txt")

    def start_log(self):
        f = open(g.app.quotas_log, "w+")
        f.close()

    def get_preference_data_from_db(self):
        """ This is just preferences - we will need to do the quotas separately """
        d = Database()
        sql = """select m.measure_sid, m.goods_nomenclature_item_id,
        m.geographical_area_id, m.ordernumber as quota_order_number_id, m.measure_type_id,
        string_agg(
            case when mc.monetary_unit_code is null then trim(to_char(mc.duty_amount, '90.00')) || '%'
            else trim(to_char(mc.duty_amount, '99990.00')) || ' GBP / ' || mc.measurement_unit_code || coalesce(mc.measurement_unit_qualifier_code, '')
            end, 
            case when mc.duty_expression_id in ('17', '35') then ' MAX '
            when mc.duty_expression_id in ('15') then ' MIN '
            else ' + ' end order by mc.duty_expression_id) as duty
        from utils.materialized_measures_real_end_dates m, measure_components mc
        where m.measure_sid = mc.measure_sid 
        and m.measure_type_id in ('142', '145')
        and m.geographical_area_id = '""" + self.origin + """'
        and (m.validity_end_date is null or m.validity_end_date >= '2022-01-01')
        group by m.measure_sid, m.goods_nomenclature_item_id, m.geographical_area_id, m.ordernumber, m.measure_type_id
        order by m.goods_nomenclature_item_id, m.measure_sid;"""

        self.db_preferences_dict = {}

        rows = d.run_query(sql.replace("\n", ""))
        for row in rows:
            measure_sid = row[0]
            goods_nomenclature_item_id = row[1]
            geographical_area_id = row[2]
            quota_order_number_id = row[3]
            measure_type_id = row[4]
            duty = self.format_duty(row[5])
            item = {
                    "goods_nomenclature_item_id": goods_nomenclature_item_id,
                    "duty": duty,
                    "geographical_area_id": geographical_area_id,
                    "quota_order_number_id": quota_order_number_id,
                    "measure_type_id": measure_type_id,
                    "measure_sid": measure_sid
                }
            if goods_nomenclature_item_id not in self.db_preferences_dict:
                self.db_preferences_dict[goods_nomenclature_item_id] = []

            self.db_preferences_dict[goods_nomenclature_item_id].append(item)

    def format_duty(self, s):
        s = s.strip()
        s = s.replace("ASVX", "% vol /  hl")
        s = s.replace("GBP/", "GBP /")
        s = s.replace("/100", "/ 100")
        s = s.replace("100kg", "100 kg")
        s = s.replace("KGMP", "kg / lactic matter")
        s = s.replace("KGMT", "kg / dry lactic matter")
        s = s.replace("DTNZ", "100 kg / %sacchar.")
        s = s.replace("DTNR", "100 kg std qual")
        s = s.replace("DTNE", "100 kg / net drained wt")
        s = s.replace("DTNM", "100 kg / net dry")
        s = s.replace("DTN", "100 kg")
        s = s.replace("TNE", "tonne")
        s = s.replace("MIL", "1000 p/st")
        s = s.replace("HLT", "hl")
        s = s.replace("kg mas", "kg / net dry")
        
        s = s.replace("(end-use)", "")
        s = s.replace("100 kg eda", "100 kg / net drained wt")
        
        s = re.sub(r"/([^ ])", r"/ \1", s)
        s = re.sub(r"([^ ])/", r"\1 /", s)
        s = re.sub(r"^([0-9])%", r"\1.00%", s)
        s = re.sub(r"^([0-9])\.([0-9])%", r"\1.\2~0%", s)
        s = re.sub(r"\.([0-9]) GBP", r".\1~0 GBP", s)
        s = s.replace("  ", " ")
        s = s.replace("~", "")
        return s.strip()
    
    def check_for_quotas(self):
        if len(self.wb.sheetnames) == 2:
            self.has_quotas = True
        else:
            self.has_quotas = False
    
    def get_preferences_from_excel(self):
        self.fta_preferences_dict = {}
        self.wb = load_workbook(filename = self.source)
        self.check_for_quotas()
        sht = self.wb.worksheets[0]
        row_count = sht.max_row
        col_count = sht.max_column
        for i in range(2, row_count + 1):
            if col_count == 4:
                preference = Preference()
                preference.goods_nomenclature_item_id = sht.cell(row=i, column=1).value
                preference.duty = sht.cell(row=i, column=2).value
                preference.staging = sht.cell(row=i, column=3).value
                preference.notes = sht.cell(row=i, column=4).value
            else:
                preference = Preference()
                preference.goods_nomenclature_item_id = sht.cell(row=i, column=1).value
                preference.duty = sht.cell(row=i, column=2).value
                
            preference.duty = self.format_duty(preference.duty)
            preference.cleanse()

            if preference.goods_nomenclature_item_id in self.fta_preferences_dict:
                self.fta_preferences_dict[preference.goods_nomenclature_item_id]["instance_count"] += 1
            else:
                self.fta_preferences_dict[preference.goods_nomenclature_item_id] = preference.as_dict()
        
        a = 1

    def compare_preferences(self):
        """ First, check that all the duties in the document are correct and present """
        for fta_duty in self.fta_preferences_dict:
            if fta_duty in self.db_preferences_dict:
                fta_duty_value = self.fta_preferences_dict[fta_duty]["duty"]
                db_duty_value = self.db_preferences_dict[fta_duty][0]["duty"]
                if fta_duty_value != db_duty_value:
                    e = PreferenceError(self.origin, self.country, "Mismatched duty", "preference", fta_duty, None, fta_duty_value, db_duty_value)
                    self.preference_errors.append(e.as_dict())
            else:
                e = PreferenceError(self.origin, self.country, "Record in document not in the database", "preference", fta_duty, None, fta_duty_value, None)
                self.preference_errors.append(e.as_dict())

        """ Second, check that all the duties in the database are reflected in the document as well """
        for db_duty in self.db_preferences_dict:
            db_duty_value = self.db_preferences_dict[db_duty][0]["duty"]
            db_duty_measure_type_id = self.db_preferences_dict[db_duty][0]["measure_type_id"]
            if db_duty not in self.fta_preferences_dict:
                e = PreferenceError(self.origin, self.country, "Record in database not in the document", "preference", db_duty, db_duty_measure_type_id, None, db_duty_value)
                self.preference_errors.append(e.as_dict())

    def write_preference_errors(self):
        with open(self.dest, 'w') as f:
            json.dump(self.preference_errors, f, indent=4)

    def check_hierarchy(self):
        # Check the items where there is a listing in the document, but not in the database
        for preference_error in self.preference_errors:
            if preference_error["commodity_code"] == "0306920000":
                a = 1
            if preference_error["reason"] == 'Record in document not in the database':
                my_descendants = preference_error["descendants"]
                compared_descendants = []
                for preference_error2 in self.preference_errors:
                    if preference_error2["commodity_code"].startswith(preference_error["significant_digits"]):
                        if preference_error2["reason"] == "Record in database not in the document":
                            compared_descendants += preference_error2["descendants"]
                            preference_error2["ignore"] = True

                compared_descendants = list(set(compared_descendants))
                my_descendants = sorted(my_descendants)
                compared_descendants = sorted(compared_descendants)
                if my_descendants != compared_descendants:
                    # This looks for occasions where there are child codes that are referenced 
                    # in a parent code (e.g. a chapter) and there is not record on the database
                    # for that code, i.e. it is missing
                    differences = []
                    for item in my_descendants:
                        if item not in compared_descendants:
                            differences.append(item)
                    
                    preference_error["differences"] = differences
                    a = 1

    def expand_differences(self):
        for preference_error in self.preference_errors:
            if len(preference_error["differences"]) > 0:
                sid_string_list = [str(int) for int in preference_error["differences"]]
                sid_string = ",".join(sid_string_list)
                
                d = Database()
                sql = """select goods_nomenclature_item_id
                from goods_nomenclatures gn
                where goods_nomenclature_sid in (""" + sid_string + """)"""
                rows = d.run_query(sql.replace("\n", ""))
                comm_code_differences = []
                for row in rows:
                    if row[0] != preference_error["commodity_code"]:
                        comm_code_differences.append(row[0])
                
                if len(comm_code_differences) == 0:
                    preference_error["ignore"] = True
                preference_error["comm_code_differences"] = sorted(comm_code_differences)
                a = 1
