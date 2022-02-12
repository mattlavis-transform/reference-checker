import json
import re
import sys
from docx import Document
from openpyxl import Workbook, load_workbook
import classes.globals as g
from classes.preference import Preference
from classes.database import Database


class PharmaChecker(object):

    def __init__(self):
        print("Checking pharma - table 10")
        self.start_log()
        self.get_pharma_from_excel()

    def start_log(self):
        f = open(g.app.pharma_log, "w+")
        f.close()

    def get_pharma_from_excel(self):
        wb = load_workbook(filename=g.app.pharma_source)
        sht = wb.active
        row_count = sht.max_row
        self.codes = {}

        min_code = "9999999999"
        max_code = "0000000000"

        for i in range(2, row_count + 1):
            cell_obj = sht.cell(row=i, column=1)
            code = cell_obj.value
            code = code.replace(" ", "")
            code = code.strip()
            code = re.sub(r"\s", "", code)
            code = code.ljust(10, "0")
            obj = {
                "sid": -1,
                "duties": [],
                "pharma_code": False,
                "zero_duty": False,
                "ancestry_array": []
            }
            self.codes[code] = obj

            if code < min_code:
                min_code = code
            if code > max_code:
                max_code = code

        # Check in the database
        sql = """
        select goods_nomenclature_item_id, goods_nomenclature_sid
        from goods_nomenclatures where validity_end_date is null
        and goods_nomenclature_item_id >= '""" + min_code + """'
        and goods_nomenclature_item_id <= '""" + max_code + """'
        and producline_suffix = '80' order by 1;
        """
        self.omissions = []
        d = Database("uk")
        real_codes = []
        self.codes_and_sids = {}
        rows = d.run_query(sql)
        for row in rows:
            real_codes.append(row[0])
            self.codes_and_sids[row[0]] = row[1]
            a = 1

        for code in self.codes:
            if code in self.codes_and_sids:
                self.codes[code]["sid"] = self.codes_and_sids[code]
            else:
                self.omissions.append(code)

        a = 1

        self.omission_count = 0
        self.matched_count = 0
        self.unmatched_count = 0

        self.unmatched = []
        self.matched = []

        if self.omissions:
            f = open(g.app.pharma_log, "w")
            for omission in self.omissions:
                f.write(omission + "\n")
                self.omission_count += 1

            f.close()

        self.check_mfn()

    def check_mfn(self):
        for code in self.codes:
            if self.codes[code] != -1:
                a = 1
                sql = "select h.goods_nomenclature_item_id, mtcd.additional_code, mtcd.duty, " \
                    "h.ancestry_array from utils.hierarchy h, utils.materialized_third_country_duties mtcd  " \
                    "where mtcd.goods_nomenclature_sid = h.goods_nomenclature_sid  " \
                    "and (%s = any(ancestry_array) or h.goods_nomenclature_sid = %s)"
                d = Database("uk")
                params = [
                    self.codes[code]["sid"],
                    self.codes[code]["sid"]
                ]
                rows = d.run_query(sql, params)
                for row in rows:
                    obj = {
                        "additional_code": row[1],
                        "duty": row[2]
                    }
                    self.codes[code]["duties"].append(obj)
                    self.codes[code]["ancestry_array"] = row[3]

                    if row[1] is None:
                        self.codes[code]["pharma_code"] = False
                        if row[2] == "0.00%":
                            self.codes[code]["zero_duty"] = True
                    else:
                        self.codes[code]["pharma_code"] = True
                        if row[2] == "0.00%":
                            self.codes[code]["zero_duty"] = True

        for code in self.codes:
            if code not in self.omissions:
                if self.codes[code]["pharma_code"] is False:
                    if self.codes[code]["zero_duty"] is False:
                        a = 1

        # Now check the ones that has 2500 and 2501 against them
        # and 
        all_sids = []
        for code in self.codes:
            if code not in self.omissions:
                all_sids.append(self.codes[code]["sid"])
                all_sids += self.codes[code]["ancestry_array"]
                a = 1
                
        all_sids = list(set(all_sids))

        sql = "select h.goods_nomenclature_sid, h.goods_nomenclature_item_id, " \
        "mtcd.additional_code, mtcd.duty, h.ancestry_array  " \
        "from utils.hierarchy h, utils.materialized_third_country_duties mtcd  " \
        "where mtcd.goods_nomenclature_sid = h.goods_nomenclature_sid  " \
        "and mtcd.additional_code in ('2500', '2501');"
        d = Database("uk")
        rows = d.run_query(sql)
        pharma_ids = []
        for row in rows:
            pharma_ids.append(row[0])
            # pharma_ids += row[4]

        pharma_ids = list(set(pharma_ids))
        excluded = []
        for pharma_id in pharma_ids:
            if pharma_id not in all_sids:
                excluded.append(pharma_id)
        a = 1
        
        excluded = list(set(excluded))
        excluded = sorted(excluded)
        excluded_string = ', '.join(str(e) for e in excluded)
        
        intersection_set = set.intersection(set(excluded), set(all_sids))
        a = 1
        
        self.excluded_comm_codes = []
        sql = "select goods_nomenclature_item_id " \
            "from utils.hierarchy h " \
            "where goods_nomenclature_sid in (" + excluded_string + ")"
        d = Database("uk")
        rows = d.run_query(sql)
        if rows:
            f = open("pharma_omissions.txt", "w")
            for row in rows:
                self.excluded_comm_codes.append(row[0])

                f.write(row[0] + "\n")
            f.close()